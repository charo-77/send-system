from openpyxl import load_workbook
from pathlib import Path
p = Path(r'D:\milu_publish_reverse_20260513\CK.xlsx')
wb = load_workbook(p)
print(wb.sheetnames)
for ws in wb.worksheets:
    print('SHEET', ws.title, ws.max_row, ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
        print(row)
